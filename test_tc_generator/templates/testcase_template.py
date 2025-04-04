from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def create_testcase_template(output_path='testcase_template.xlsx'):
    """
    Create an Excel template for testcases with predefined format
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Testcases"
    
    # Define column headers
    headers = [
        "항목번호", "대분류", "중분류", "소분류", "구분", 
        "테스트 내용", "테스트 조건", "기대 결과", "비고", "점수", "등급"
    ]
    
    # Set column widths
    column_widths = {
        'A': 10,  # 항목번호
        'B': 15,  # 대분류
        'C': 15,  # 중분류
        'D': 15,  # 소분류
        'E': 10,  # 구분
        'F': 40,  # 테스트 내용
        'G': 30,  # 테스트 조건
        'H': 30,  # 기대 결과
        'I': 15,  # 비고
        'J': 10,  # 점수
        'K': 10,  # 등급
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set styles
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Create thin border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Add example row
    example_data = [
        1,
        "시스템",
        "로그인",
        "오류 메시지",
        "예외",
        "로그인 실패 메시지 확인",
        "잘못된 ID/PW 입력 시",
        "오류 메시지 정확히 출력",
        "-",
        95,
        "🟢"
    ]
    
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(vertical='center')
        cell.border = thin_border
        
        # Center specific columns
        if col_idx in [1, 5, 10, 11]:
            cell.alignment = center_alignment
    
    # Add 10 empty rows with formatting
    for row_idx in range(3, 13):
        # Add row number in first column
        cell = ws.cell(row=row_idx, column=1, value=row_idx-1)
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Format remaining cells
        for col_idx in range(2, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.border = thin_border
            
            # Center specific columns
            if col_idx in [1, 5, 10, 11]:
                cell.alignment = center_alignment
    
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Save workbook
    wb.save(output_path)
    print(f"Template created: {output_path}")
    
    return output_path

if __name__ == "__main__":
    # Create template in the templates directory
    template_path = os.path.join(os.path.dirname(__file__), "testcase_template.xlsx")
    create_testcase_template(template_path)