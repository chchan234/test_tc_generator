import streamlit as st
import os
import tempfile
from langchain.schema import Document
from langchain_community.document_loaders import PyMuPDFLoader, UnstructuredWordDocumentLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO

from processor import (
    filter_unnecessary_sentences,
    identify_document_structure,
    generate_testcases,
    validate_testcase_quality
)

st.set_page_config(page_title="게임 기획서 → Testcase 자동 생성기", layout="wide")

def save_uploaded_file(uploaded_file):
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as temp_file:
        temp_file.write(uploaded_file.getvalue())
        return temp_file.name

def load_document(file_path):
    file_extension = os.path.splitext(file_path)[1].lower()
    
    if file_extension == '.pdf':
        loader = PyMuPDFLoader(file_path)
        documents = loader.load()
    elif file_extension in ['.docx', '.doc']:
        loader = UnstructuredWordDocumentLoader(file_path)
        documents = loader.load()
    else:
        st.error("지원하지 않는 파일 형식입니다. PDF 또는 DOCX 파일만 업로드해주세요.")
        return None
    
    return documents

def split_into_sentences(documents):
    text_splitter = RecursiveCharacterTextSplitter(
        separators=["\n\n", "\n", ".", "!", "?"],
        chunk_size=1000,
        chunk_overlap=0
    )
    
    sentences = []
    for doc in documents:
        chunks = text_splitter.split_text(doc.page_content)
        for chunk in chunks:
            sentences.append(Document(page_content=chunk, metadata=doc.metadata))
    
    return sentences

def create_excel_with_testcases(testcases):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Testcases"
    
    # Define column headers
    headers = ["항목번호", "대분류", "중분류", "소분류", "구분", "테스트 내용", "테스트 조건", "기대 결과", "비고", "점수", "등급"]
    
    # Set header style
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_font = Font(bold=True)
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write testcase data
    for row_idx, tc in enumerate(testcases, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1)  # 항목번호
        ws.cell(row=row_idx, column=2, value=tc["대분류"])
        ws.cell(row=row_idx, column=3, value=tc["중분류"])
        ws.cell(row=row_idx, column=4, value=tc["소분류"])
        ws.cell(row=row_idx, column=5, value=tc["구분"])
        ws.cell(row=row_idx, column=6, value=tc["테스트 내용"])
        ws.cell(row=row_idx, column=7, value=tc["테스트 조건"])
        ws.cell(row=row_idx, column=8, value=tc["기대 결과"])
        ws.cell(row=row_idx, column=9, value=tc["비고"])
        
        # Add score and grade with conditional formatting
        score = tc["점수"]
        ws.cell(row=row_idx, column=10, value=f"{score}점")
        
        # Set grade based on score
        if score >= 90:
            grade = "🟢"
        elif score >= 70:
            grade = "🟡"
        elif score >= 50:
            grade = "🟠"
        else:
            grade = "🔴"
        
        ws.cell(row=row_idx, column=11, value=grade)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO object
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def main():
    st.title("게임 기획서 → Testcase 자동 생성기")
    st.write("게임 기획서를 업로드하면 AI가 자동으로 testcase를 생성하고 품질을 검증합니다.")
    
    # Sidebar for model selection and API key input
    st.sidebar.header("⚙️ 설정")
    
    model_option = st.sidebar.selectbox(
        "AI 모델 선택",
        ["Google Gemini", "OpenAI GPT-4 Turbo"]
    )
    
    api_keys = {}
    
    if model_option == "Google Gemini":
        gemini_api = st.sidebar.text_input("Google Gemini API 키 입력", type="password")
        if gemini_api:
            api_keys["gemini"] = gemini_api
    else:
        openai_api = st.sidebar.text_input("OpenAI API 키 입력", type="password")
        if openai_api:
            api_keys["openai"] = openai_api
    
    # File uploader
    uploaded_file = st.file_uploader("기획서(DOCX, PDF)를 업로드해주세요", type=['docx', 'pdf'])
    
    if uploaded_file and api_keys:
        with st.spinner("문서 분석 중입니다..."):
            # Save uploaded file
            temp_file_path = save_uploaded_file(uploaded_file)
            
            # Load document
            documents = load_document(temp_file_path)
            if not documents:
                return
            
            # Split into sentences
            st.info("문서를 문장 단위로 분할합니다...")
            sentences = split_into_sentences(documents)
            
            # Filter unnecessary sentences
            st.info("Testcase에 필요한 문장을 필터링합니다...")
            filtered_sentences = filter_unnecessary_sentences(sentences, model_option, api_keys)
            
            # Identify document structure
            st.info("문서 구조를 분석하여 대/중/소분류를 식별합니다...")
            doc_structure = identify_document_structure(filtered_sentences, model_option, api_keys)
            
            # Generate testcases
            st.info("Testcase를 생성합니다...")
            testcases = generate_testcases(filtered_sentences, doc_structure, model_option, api_keys)
            
            # Validate testcase quality
            st.info("생성된 Testcase의 품질을 검증합니다...")
            validated_testcases = validate_testcase_quality(testcases, model_option, api_keys)
            
            # Create Excel file
            excel_file = create_excel_with_testcases(validated_testcases)
            
            # Display success message and download button
            st.success("Testcase가 성공적으로 생성되었습니다!")
            
            # Display testcase preview
            st.subheader("Testcase 미리보기")
            preview_df = pd.DataFrame(validated_testcases)
            st.dataframe(preview_df)
            
            # Download button
            st.download_button(
                label="Testcase 엑셀 파일 다운로드",
                data=excel_file,
                file_name="generated_testcases.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Clean up temporary file
            os.unlink(temp_file_path)
    
    else:
        if not api_keys and uploaded_file:
            st.warning("API 키를 입력해주세요.")
        elif not uploaded_file and api_keys:
            st.warning("기획서 파일을 업로드해주세요.")

if __name__ == "__main__":
    main()
